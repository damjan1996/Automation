import requests
import time
import openpyxl
from datetime import datetime

# Open the Excel spreadsheet
workbook = openpyxl.load_workbook('03 - product_data_upc.xlsx')
sheet = workbook.active

# API endpoint (the server URL where the API requests are being sent)
base_url = 'https://[your_business].app.apparelmagic.com/api/json/inventory'

# Authentication token
token_value = 'your_authentication_token_here'

# Initialize counter for updated products
updated_products_count = 0

# Iterate over the spreadsheet rows, starting from row 2 to skip the header
for row in range(2, sheet.max_row + 1):
    product_id = sheet.cell(row=row, column=1).value  # Assuming SKU ID is in Column A
    gtin = sheet.cell(row=row, column=19).value  # Assuming GTIN is in Column S

    # Construct the URL for this specific product by appending the product_id to the base_url
    url = base_url + str(product_id)

    # Calculate current time for every request
    time_value = str(int(time.time()))

    # Define authentication parameters
    auth_params = {
        "time": time_value,
        "token": token_value
    }

    # Define the data to be updated (GTIN value in this case)
    payload = {
        "sku_id": product_id,
        "upc_display": gtin
    }

    # Add authentication info to the payload for the PUT request
    payload.update(auth_params)

    # Send the updated data to the API (authentication info is not in the URL)
    update_response = requests.put(url, json=payload)

    # Check if the update was successful
    if update_response.status_code == 200:
        updated_products_count += 1
        print(f"Products updated so far: {updated_products_count}")
        print(f"Successfully assigned GTIN {gtin} to product {product_id}.")
    else:
        print(f"Error updating product {product_id} with GTIN {gtin}")
