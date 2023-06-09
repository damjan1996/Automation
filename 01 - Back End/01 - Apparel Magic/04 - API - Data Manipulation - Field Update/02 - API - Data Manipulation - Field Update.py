import requests
import time
import openpyxl
from datetime import datetime

# Load the Excel workbook
workbook = openpyxl.load_workbook('your_file_name.xlsx')
sheet = workbook.active

# API endpoint
base_url = 'https://[your_business].app.apparelmagic.com/api/json/inventory'

# Authentication token - you must use your actual authentication token
token_value = 'your_authentication_token_here'

# Initialize the count of updated products
updated_products_count = 0

# Iterate over the rows in the table (starting from row 2 to skip the header)
for row in range(2, sheet.max_row + 1):
    # Assume the product ID is in column A
    product_id = sheet.cell(row=row, column=1).value

    # Construct the URL by appending the product ID to the base URL
    url = base_url + str(product_id)

    # Generate the current time for each request
    time_value = str(int(time.time()))

    # Define authentication parameters
    auth_params = {
        "time": time_value,
        "token": token_value
    }

    # Define the product attributes to be updated
    payload = {
        "product_tags": sheet.cell(row=row, column=10).value,  # Assume the product tags are in column 73
        "product_category": sheet.cell(row=row, column=11).value,  # Assume the product category is in column 74
        "product_shop_color": sheet.cell(row=row, column=12).value,  # Assume the product color is in column 75
        "product_shop_by_style": sheet.cell(row=row, column=13).value,  # Assume the product style is in column 76
    }

    # Add the authentication parameters to the request payload
    payload.update(auth_params)

    # Send the updated data to the API (excluding the authentication parameters from the URL)
    update_response = requests.put(url, json=payload)

    # Check if the update was successful
    if update_response.status_code == 200:
        updated_products_count += 1
        print(f"Updated products so far: {updated_products_count}")
        print(f"The properties of product {product_id} were successfully updated.")
    else:
        print(f"Error updating product {product_id}")
